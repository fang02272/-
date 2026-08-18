# -*- coding: utf-8 -*-
"""导入卡诺普机器人焊接工艺图谱 Excel → 真值库 weld_cases.json（全部25个sheet）

用法:
  python tools/import_kanopu.py <Excel路径>
  例: python tools/import_kanopu.py "C:/xxx/机器人焊接工艺图谱生成(实际电压11.14).xlsx"

Excel 结构（26 sheet）：
- 参数说明: 数据说明（基准参数 + 只变一个参数的可焊接范围）
- 主参数 sheet: 碳钢中厚板/薄板/镀锌板/不锈钢 → 基准准确参数
- 变化 sheet: 电流小/大、电压小/大、速度快/慢、焊枪角度范围 → 范围边界

每条记录含 variant 标签（基准/电流小/电流大/电压小/电压大/速度快/速度慢），
用于工艺卡片按需选基准或范围。
"""
import json
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# sheet 名 → (材料, 板厚字段, variant 标签)
SHEET_MAP = [
    ("碳钢中厚板参数", "碳钢", "焊脚", "基准"),
    ("碳钢薄板参数", "碳钢", "板厚", "基准"),
    ("碳钢薄板电流小", "碳钢", "板厚", "电流小"),
    ("碳钢薄板电流大", "碳钢", "板厚", "电流大"),
    ("碳钢薄板速度慢", "碳钢", "板厚", "速度慢"),
    ("碳钢薄板速度快", "碳钢", "板厚", "速度快"),
    ("碳钢薄板电压大", "碳钢", "板厚", "电压大"),
    ("碳钢薄板电压小", "碳钢", "板厚", "电压小"),
    ("碳钢薄板焊枪角度范围", "碳钢", "板厚", "角度范围"),
    ("镀锌板参数", "镀锌板", "板厚", "基准"),
    ("镀锌板电流小", "镀锌板", "板厚", "电流小"),
    ("镀锌板电流大", "镀锌板", "板厚", "电流大"),
    ("镀锌板速度慢", "镀锌板", "板厚", "速度慢"),
    ("镀锌板速度快", "镀锌板", "板厚", "速度快"),
    ("镀锌板电压大", "镀锌板", "板厚", "电压大"),
    ("镀锌板电压小", "镀锌板", "板厚", "电压小"),
    ("镀锌板焊枪角度范围", "镀锌板", "板厚", "角度范围"),
    ("不锈钢参数 ", "不锈钢", "板厚", "基准"),
    ("不锈钢电流小", "不锈钢", "板厚", "电流小"),
    ("不锈钢电流大", "不锈钢", "板厚", "电流大"),
    ("不锈钢速度慢", "不锈钢", "板厚", "速度慢"),
    ("不锈钢速度快", "不锈钢", "板厚", "速度快"),
    ("不锈钢电压大", "不锈钢", "板厚", "电压大"),
    ("不锈钢电压小", "不锈钢", "板厚", "电压小"),
    ("不锈钢焊枪角度范围", "不锈钢", "板厚", "角度范围"),
]


def parse_angle(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_num(v):
    """数字：'<60' → 60；190 → 190；None → None"""
    if v is None:
        return None
    s = str(v).strip()
    m = re.search(r'\d+(\.\d+)?', s)
    return float(m.group(0)) if m else None


def main():
    if len(sys.argv) < 2:
        print("用法: python tools/import_kanopu.py <Excel路径>")
        sys.exit(1)
    xlsx = sys.argv[1]
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    cases = []
    seen = set()

    for sheet_name, material, thickness_key, variant in SHEET_MAP:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️ 缺 sheet: {sheet_name}")
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        # 找表头行（含 '材料'）
        header_idx = None
        for i, row in enumerate(rows[:5]):
            if row and row[0] == '材料':
                header_idx = i
                break
        if header_idx is None:
            continue
        headers = [str(h).strip() if h else "" for h in rows[header_idx]]
        # 字段索引（电流/电压/速度 可能在重复列，取第一个）
        col_idx = {}
        for name in ["材料", "气体", "焊丝直径", thickness_key, "焊缝形式", "电流", "电压", "焊接速度", "干伸长"]:
            if name in headers:
                col_idx[name] = headers.index(name)
        if "材料" not in col_idx:
            continue

        # 焊枪角度：表头"焊枪角度"列 = 前后，+1 = 左右（单位行标注）
        angle_fb_col = headers.index("焊枪角度") if "焊枪角度" in headers else None
        angle_lr_col = angle_fb_col + 1 if angle_fb_col is not None else None
        # 摆幅/频率/停留（中厚板有）
        weave_col = headers.index("摆幅（mm）") if "摆幅（mm）" in headers else None
        freq_col = headers.index("频率") if "频率" in headers else None
        dwell_col = headers.index("停留时间（s）") if "停留时间（s）" in headers else None

        for row in rows[header_idx + 2:]:
            if not row or not row[col_idx["材料"]]:
                continue
            m_name = str(row[col_idx["材料"]]).strip()
            if m_name in ("材料", ""):
                continue
            case = {
                "material": material,
                "gas": str(row[col_idx["气体"]]).strip() if "气体" in col_idx and row[col_idx["气体"]] else "80%Ar+20%CO2",
                "wire_dia": str(row[col_idx["焊丝直径"]]).strip() if "焊丝直径" in col_idx and row[col_idx["焊丝直径"]] else "1.2mm",
                "thickness": parse_num(row[col_idx[thickness_key]]) if thickness_key in col_idx else None,
                "joint": str(row[col_idx["焊缝形式"]]).strip() if "焊缝形式" in col_idx and row[col_idx["焊缝形式"]] else "",
                "current": parse_num(row[col_idx["电流"]]) if "电流" in col_idx else None,
                "voltage": parse_num(row[col_idx["电压"]]) if "电压" in col_idx else None,
                "speed": parse_num(row[col_idx["焊接速度"]]) if "焊接速度" in col_idx else None,
                "stick_out": parse_num(row[col_idx["干伸长"]]) if "干伸长" in col_idx else 15,
                "gun_angle_fb": parse_angle(row[angle_fb_col]) if angle_fb_col is not None and len(row) > angle_fb_col else None,
                "gun_angle_lr": parse_angle(row[angle_lr_col]) if angle_lr_col is not None and len(row) > angle_lr_col else None,
                "variant": variant,
            }
            if weave_col is not None and len(row) > weave_col:
                case["weave_width"] = parse_num(row[weave_col])
            if freq_col is not None and len(row) > freq_col:
                case["weave_freq"] = parse_num(row[freq_col])
            if dwell_col is not None and len(row) > dwell_col:
                case["weave_dwell"] = parse_num(row[dwell_col])
            # 去重（材料+厚度+焊缝形式+variant）
            key = (case["material"], case["thickness"], case["joint"], case["variant"])
            if key in seen:
                continue
            seen.add(key)
            cases.append(case)

    # 保存
    out = PROJECT_ROOT / "data" / "weld_cases.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(cases, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"✅ 导入 {len(cases)} 条 → {out}")
    # 统计
    from collections import Counter
    mats = Counter(c["material"] for c in cases)
    print(f"   材料分布: {dict(mats)}")
    variants = Counter(c["variant"] for c in cases)
    print(f"   variant分布: {dict(variants)}")
    # 按材料+板厚抽样
    for m in mats:
        tc = Counter(c["thickness"] for c in cases if c["material"] == m and c["variant"] == "基准")
        print(f"   {m} 基准: 板厚档 {sorted(tc.keys())}")


if __name__ == "__main__":
    main()
