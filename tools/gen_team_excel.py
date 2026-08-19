# -*- coding: utf-8 -*-
"""生成团队资源 Excel（本地 docs/，不进 git）

Sheet:
1. 数据缺口资源   — 铝合金/铸铁/管道6G/问答对 的公开资源+搜索建议
2. 工艺真值模板   — Excel 列模板（喂数据用）
3. 问答对模板     — JSON 模板（测试集/专家库用）
4. 管道6G模板     — 分段焊接模板
5. 团队任务       — 4人八天任务
6. 验收指标       — 下周五目标
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# 样式
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_sheet(ws, headers, rows, widths, sub_rows=None):
    """写表头+数据，带样式"""
    # 表头
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
    # 数据
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(ri, ci, v)
            c.border = BORDER
            c.alignment = WRAP
            # 子分类行高亮
            if sub_rows and ri in sub_rows.get(ci, set()):
                c.fill = SUB_FILL
    # 列宽
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


wb = Workbook()

# ============ Sheet1: 数据缺口资源 ============
ws1 = wb.active
ws1.title = "数据缺口资源"
headers1 = ["待补数据", "可参考资源", "说明", "可提取内容", "优先级", "负责人"]
rows1 = [
    ["铝合金真值", "学术论文：铝合金焊接共享数据库系统研究", "专门研究铝合金工艺参数+材料匹配", "电流/电压/焊速/焊丝牌号/保护气", "重要", "徐采妮"],
    ["铝合金真值", "AWS 铝合金标准焊接工艺规程（SWPS）草案", "权威标准工艺规程", "工艺参数+坡口+焊材", "重要", "徐采妮"],
    ["铝合金真值", "GB/T 10858 铝及铝合金焊丝", "焊丝牌号标准", "焊丝型号", "辅助", "徐采妮"],
    ["铝合金真值", "期刊《焊接学报》《中国有色金属学报》", "铝合金焊接工艺参数论文", "具体参数组合", "辅助", "徐采妮"],
    ["铸铁真值", "GB/T 44841-2024《非合金及低合金铸铁焊接工艺评定试验》★", "最权威直接（2025.5实施）", "预热/电流/焊条牌号/工艺", "最高优先", "徐采妮"],
    ["铸铁真值", "GB/T 10044 铸铁焊条", "焊条标准", "焊条型号", "辅助", "徐采妮"],
    ["铸铁真值", "Z308/Z408 焊条厂家工艺参数", "厂家推荐参数", "电流/预热", "辅助", "徐采妮"],
    ["管道6G参数", "海底管线6G位置焊接参数专利", "专利详述参数", "分段角度/电流电压", "重要", "徐采妮+吴杰"],
    ["管道6G参数", "6G位置不锈钢管GTAW焊接工艺论文", "学术研究", "工艺参数+操作要点", "重要", "徐采妮+吴杰"],
    ["管道6G参数", "6G位置GTAW操作方法综述文章", "系统介绍操作与参数", "操作方法+参数", "重要", "徐采妮+吴杰"],
    ["管道6G参数", "ISO 9606-1 / API 1104", "管道焊接标准", "焊接要求/验收", "辅助", "徐采妮+吴杰"],
    ["真实问答对", "论文/专利的'背景技术'与'具体实施方式'", "隐含工程问题+解决方案", "问答对（Q现象/A方案）", "重要", "曾立夫"],
    ["真实问答对", "焊接论坛/社区高频问题", "焊缝缺陷/参数选择", "真实用户问题", "重要", "曾立夫"],
    ["真实问答对", "标准条文的'为什么'", "预热为什么/探伤为什么", "概念问答对", "辅助", "曾立夫"],
]
style_sheet(ws1, headers1, rows1, [14, 44, 30, 28, 12, 12])

# ============ Sheet2: 工艺真值模板 ============
ws2 = wb.create_sheet("工艺真值模板")
headers2 = ["材料", "气体", "焊丝直径", "板厚/焊脚", "焊缝形式", "电流", "电压",
            "焊接速度", "干伸长", "焊枪角度(前后)", "焊枪角度(左右)", "variant", "来源"]
rows2 = [
    ["铝合金", "纯Ar", "1.2mm", "3mm", "平拼接", 110, 19.5, 35, 15, 80, 90, "基准", "GB/T+论文"],
    ["铸铁", "—", "Z408", "10mm", "平拼接", 130, 22, 15, 15, 80, 90, "基准", "GB/T 44841"],
    ["碳钢", "80%Ar+20%CO2", "1.2mm", "5mm", "船型", 250, 23.5, 8, 15, 80, 90, "基准", "卡诺普实测"],
    ["碳钢", "80%Ar+20%CO2", "1.2mm", "5mm", "船型", 280, 24.5, 8, 15, 80, 90, "电流大", "卡诺普实测"],
]
note2 = [["说明：variant 取 基准/电流小/电流大/电压小/电压大/速度快/速度慢/角度范围；来源写标准号/论文名，可追溯"]]
style_sheet(ws2, headers2, rows2, [10, 14, 10, 12, 12, 8, 8, 10, 8, 14, 14, 10, 18])
# 注释行
ws2.cell(6, 1, note2[0][0]).font = Font(italic=True, color="666666")

# ============ Sheet3: 问答对模板 ============
ws3 = wb.create_sheet("问答对模板")
headers3 = ["字段", "示例", "说明"]
rows3 = [
    ["q", "铝合金焊接为什么容易产生气孔？", "问题"],
    ["a", "氢在液态铝中溶解度远高于固态，凝固时析出形成气孔；需清理氧化膜、控制保护气", "答案"],
    ["topic", "铝合金缺陷", "主题分类"],
    ["source", "GB/T 44841 / 论文", "来源"],
    ["level", "概念", "概念/参数/缺陷/工艺"],
]
style_sheet(ws3, headers3, rows3, [10, 60, 20])
ws3.cell(7, 1, "JSON格式：{\"q\":\"...\",\"a\":\"...\",\"topic\":\"...\",\"source\":\"...\",\"level\":\"...\"}").font = Font(italic=True, color="666666")

# ============ Sheet4: 管道6G模板 ============
ws4 = wb.create_sheet("管道6G模板")
headers4 = ["material", "pipe_dia", "position", "seg", "current", "voltage", "angle", "source"]
rows4 = [
    ["碳钢", "100mm", "6G", "6点-3点", "90-100A", "10-12V", "15°", "API 1104"],
    ["碳钢", "100mm", "6G", "3点-12点", "80-95A", "10-11V", "5°", "API 1104"],
    ["不锈钢", "50mm", "6G", "6点-3点", "75-85A", "9-10V", "15°", "论文"],
]
style_sheet(ws4, headers4, rows4, [10, 12, 10, 12, 12, 10, 10, 14])

# ============ Sheet5: 团队任务 ============
ws5 = wb.create_sheet("团队任务")
headers5 = ["成员", "角色", "任务", "目标", "交付"]
rows5 = [
    ["赵剑乔", "检索/推理", "LLM 推理瘦身", "参数问题本地直出 +20%", "本地命中率对比"],
    ["赵剑乔", "检索/推理", "意图路由优化", "参数/概念误判 <5%", "误判率报告"],
    ["赵剑乔", "检索/推理", "工艺卡片质量", "铝合金/铸铁卡片可读", "兜底规则优化"],
    ["赵剑乔", "检索/推理", "语义向量调优", "近义检索 +10%", "检索命中对比"],
    ["吴杰", "性能", "响应时间优化", "<3s", "性能基线对比"],
    ["吴杰", "性能", "缓存命中提升", ">30%", "缓存策略"],
    ["吴杰", "性能", "索引加载优化", "语义向量 <10s", "加载策略"],
    ["吴杰", "性能", "管道6G数据支持", "与徐采妮协作", "管道卡片性能"],
    ["徐采妮", "数据", "铸铁真值（GB/T 44841）", "补 100+ 条", "铸铁真值库"],
    ["徐采妮", "数据", "铝合金真值", "补 100+ 条", "铝合金真值库"],
    ["徐采妮", "数据", "管道 6G 参数", "补 30+ 条分段", "管道卡片"],
    ["徐采妮", "数据", "焊接结构原理表格", "9→20+", "表格重建"],
    ["徐采妮", "数据", "数据模板落地", "3 个模板", "模板文档"],
    ["曾立夫", "测试", "真实问答对", "收集 50+ 条", "测试集扩充"],
    ["曾立夫", "测试", "新真值回归", "铝合金/铸铁入库验证", "回归报告"],
    ["曾立夫", "测试", "数据质量检查", "OCR/表格错位", "质量报告"],
    ["曾立夫", "测试", "文档同步", "周报/README", "文档更新"],
]
style_sheet(ws5, headers5, rows5, [12, 12, 30, 28, 20])

# ============ Sheet6: 验收指标 ============
ws6 = wb.create_sheet("验收指标")
headers6 = ["指标", "目标", "负责人", "验证方式"]
rows6 = [
    ["本地直出命中率", "≥80%", "赵剑乔", "perf_baseline + 日志"],
    ["响应时间", "<3s", "吴杰", "perf_baseline"],
    ["检索命中率", "≥90%", "赵剑乔", "测试集"],
    ["测试套件", "9 组全 PASS", "曾立夫", "tools/tests.py"],
    ["真值库", "≥1600 条", "徐采妮", "weld_cases.json 统计"],
    ["管道 6G 卡片", "可用", "徐采妮+吴杰", "问'管道6G焊'验证"],
]
style_sheet(ws6, headers6, rows6, [22, 14, 18, 24])

# 保存
out = PROJECT_ROOT / "docs" / "团队资源汇总.xlsx"
wb.save(out)
print(f"✅ 已生成: {out}")
